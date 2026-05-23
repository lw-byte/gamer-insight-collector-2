#!/usr/bin/env python3
"""
荣耀建议广场数据抓取脚本 v2
批量抓取所有页面并保存为JSON、XLSX和HTML格式
"""

import json
import re
import os
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict, field

# 导入Excel支持
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

@dataclass
class Comment:
    """评论数据模型"""
    user_id: str
    content: str
    post_time: str
    vote_count: int = 0

@dataclass
class Post:
    """帖子数据模型"""
    user_id: str = ""
    product_name: str = "荣耀WIN系列"
    title: str = ""
    content: str = ""
    post_time: str = ""
    view_count: int = 0
    comment_count: int = 0
    post_url: str = ""
    page_number: int = 1
    vote_count: int = 0  # 建议投票参与人数
    comments: List[Dict] = field(default_factory=list)
    image_urls: List[str] = field(default_factory=list)


def parse_page_structure(raw_output: str, page_num: int) -> List[Post]:
    """从firecrawl输出的原始文本解析帖子数据"""
    posts = []

    # 按换行符分割
    lines = raw_output.split('\n')

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 检测帖子标题（后面有支持/反对）
        title_match = re.match(r'^- link "([^"]+)" \[ref=e(\d+)\]$', line)
        if title_match and i+1 < len(lines) and '支持' in lines[i+1]:
            title = title_match.group(1)

            # 查找用户ID和评论数
            user_id = ''
            comment_count = 0

            # 在接下来几行找用户ID（非数字的link）和评论数（纯数字的link）
            for j in range(i+1, min(i+20, len(lines))):
                next_line = lines[j].strip()

                # 用户ID: link "用户名" [ref=eXXX]，用户名不是纯数字
                user_match = re.match(r'^- link "([^"]+)" \[ref=e(\d+)\]$', next_line)
                if user_match:
                    potential_text = user_match.group(1)
                    # 如果不是纯数字，认为是用户ID
                    if potential_text and not potential_text.isdigit():
                        user_id = potential_text
                        # 在此之后找评论数
                        for k in range(j+1, min(j+5, len(lines))):
                            count_match = re.match(r'^- link "(\d+)" \[ref=e(\d+)\]$', lines[k].strip())
                            if count_match:
                                comment_count = int(count_match.group(1))
                                break
                        break

            post = Post(
                title=title.strip(),
                user_id=user_id.strip(),
                comment_count=comment_count,
                page_number=page_num
            )
            posts.append(post)

        i += 1

    return posts


def save_json(posts: List[Post], filepath: Path):
    """保存为JSON格式"""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump([asdict(p) for p in posts], f, ensure_ascii=False, indent=2)
    print(f"  JSON已保存: {filepath}")


def save_xlsx(posts: List[Post], filepath: Path):
    """保存为XLSX格式"""
    if not OPENPYXL_AVAILABLE:
        print("  openpyxl未安装，跳过XLSX生成")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "荣耀建议广场数据"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="CF1F25", end_color="CF1F25", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        "序号", "用户ID", "产品名", "发帖标题", "发帖内容",
        "发帖时间", "浏览数量", "评论数量", "发帖链接",
        "页号", "投票参与人数", "评论详情", "图片链接"
    ]
    ws.append(headers)

    # 设置表头样式
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    for i, post in enumerate(posts, 1):
        comments_json = json.dumps(post.comments, ensure_ascii=False)
        images_json = json.dumps(post.image_urls, ensure_ascii=False)

        row = [
            i,
            post.user_id,
            post.product_name,
            post.title,
            post.content,
            post.post_time,
            post.view_count,
            post.comment_count,
            post.post_url,
            post.page_number,
            post.vote_count,
            comments_json,
            images_json
        ]
        ws.append(row)

        # 设置数据行样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i + 1, column=col)
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal='center')

    # 调整列宽
    column_widths = [8, 20, 15, 40, 50, 15, 10, 10, 50, 8, 12, 60, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 设置行高
    ws.row_dimensions[1].height = 25
    for i in range(2, len(posts) + 2):
        ws.row_dimensions[i].height = 40

    wb.save(filepath)
    print(f"  XLSX已保存: {filepath}")


def save_html(posts: List[Post], filepath: Path):
    """保存为HTML格式"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>荣耀建议广场 - 游戏体验板块 - 荣耀WIN系列数据</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #cf1f25;
            text-align: center;
            margin-bottom: 10px;
        }}
        .subtitle {{
            text-align: center;
            color: #666;
            margin-bottom: 30px;
        }}
        .stats {{
            display: flex;
            justify-content: space-around;
            margin-bottom: 30px;
            padding: 20px;
            background: #f9f9f9;
            border-radius: 8px;
        }}
        .stat-item {{
            text-align: center;
        }}
        .stat-value {{
            font-size: 32px;
            font-weight: bold;
            color: #cf1f25;
        }}
        .stat-label {{
            color: #666;
            font-size: 14px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 12px 8px;
            text-align: left;
            vertical-align: top;
        }}
        th {{
            background-color: #cf1f25;
            color: white;
            font-weight: 600;
            position: sticky;
            top: 0;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f0f0f0;
        }}
        .title-cell {{
            max-width: 300px;
            word-wrap: break-word;
        }}
        .comment-count {{
            text-align: center;
            font-weight: bold;
        }}
        .view-count {{
            text-align: center;
            color: #666;
        }}
        .page-num {{
            text-align: center;
        }}
        .link-cell a {{
            color: #0066cc;
            text-decoration: none;
        }}
        .link-cell a:hover {{
            text-decoration: underline;
        }}
        .timestamp {{
            text-align: center;
            color: #999;
            margin-top: 20px;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📱 荣耀建议广场 - 游戏体验板块</h1>
        <p class="subtitle">荣耀WIN系列用户反馈数据</p>

        <div class="stats">
            <div class="stat-item">
                <div class="stat-value">{len(posts)}</div>
                <div class="stat-label">总帖子数</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{sum(p.view_count for p in posts)}</div>
                <div class="stat-label">总浏览数</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{sum(p.comment_count for p in posts)}</div>
                <div class="stat-label">总评论数</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{len(set(p.user_id for p in posts))}</div>
                <div class="stat-label">参与用户数</div>
            </div>
        </div>

        <table>
            <thead>
                <tr>
                    <th>序号</th>
                    <th>用户ID</th>
                    <th>产品名</th>
                    <th>发帖标题</th>
                    <th>发帖时间</th>
                    <th>浏览数</th>
                    <th>评论数</th>
                    <th>页号</th>
                    <th>发帖链接</th>
                </tr>
            </thead>
            <tbody>
"""

    for i, post in enumerate(posts, 1):
        link = f'<a href="{post.post_url}" target="_blank">查看详情</a>' if post.post_url else '-'
        html += f"""                <tr>
                    <td style="text-align:center">{i}</td>
                    <td>{post.user_id}</td>
                    <td>{post.product_name}</td>
                    <td class="title-cell">{post.title}</td>
                    <td>{post.post_time or '-'}</td>
                    <td class="view-count">{post.view_count}</td>
                    <td class="comment-count">{post.comment_count}</td>
                    <td class="page-num">{post.page_number}</td>
                    <td class="link-cell">{link}</td>
                </tr>
"""

    html += f"""            </tbody>
        </table>
        <p class="timestamp">数据抓取时间: {timestamp}</p>
    </div>
</body>
</html>"""

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  HTML已保存: {filepath}")


def main():
    print("=" * 60)
    print("荣耀建议广场 - 游戏体验板块 - 荣耀WIN系列数据抓取")
    print("=" * 60)

    # 创建输出目录
    output_dir = Path("honor_data")
    output_dir.mkdir(exist_ok=True)

    # 从之前抓取的原始数据解析帖子
    raw_file = Path("honor_data/page1_raw.json")
    if raw_file.exists():
        with open(raw_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 使用result字段（包含DOM结构）
        result = data.get('result', '')
        posts = parse_page_structure(result, 1)

        print(f"\n从第1页解析到 {len(posts)} 个帖子")

        # 保存数据
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        save_json(posts, output_dir / f"honor_win_posts_{timestamp}.json")
        save_xlsx(posts, output_dir / f"honor_win_posts_{timestamp}.xlsx")
        save_html(posts, output_dir / f"honor_win_posts_{timestamp}.html")

        print(f"\nDone! Data parsing complete!")
        print(f"Output directory: {output_dir.absolute()}")
    else:
        print(f"错误: 找不到原始数据文件 {raw_file}")


if __name__ == "__main__":
    main()