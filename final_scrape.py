#!/usr/bin/env python3
"""
荣耀建议广场 - 完整数据抓取脚本
抓取所有页面数据并保存为JSON、XLSX和HTML格式
"""

import json
import re
import os
import subprocess
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from dataclasses import dataclass, field

@dataclass
class Post:
    """帖子数据模型"""
    user_id: str = ""
    product_name: str = "荣耀WIN系列"
    title: str = ""
    content: str = ""
    post_time: str = ""
    view_count: int = 0
    comment_count: int = 0
    post_url: str = ""
    page_number: int = 1
    vote_count: int = 0  # 建议投票参与人数
    comments: List[Dict] = field(default_factory=list)
    image_urls: List[str] = field(default_factory=list)


def parse_page_from_autocli(html_content: str, page_num: int) -> List[Dict]:
    """解析页面内容"""
    posts = []

    # 分割每个帖子
    thread_pattern = r'<div id="(\d+)">(.*?)</div>\s*</div>\s*<div id="'
    matches = re.findall(thread_pattern, html_content, re.DOTALL)

    for post_id, post_content in matches:
        # 提取标题
        title_match = re.search(r'<p>(.*?)</p>', post_content, re.DOTALL)
        title = title_match.group(1) if title_match else ''

        # 提取参与人数
        participation_match = re.search(r'<span>(\d+)</span>\s*人已参与', post_content)
        participation = int(participation_match.group(1)) if participation_match else 0

        # 清理HTML标签
        title = re.sub(r'<[^>]+>', '', title).strip()

        if title:
            posts.append({
                'post_id': post_id,
                'title': title,
                'vote_count': participation,
                'url': f'https://club.honor.com/cn//cn//opinion_thread-view.html?id={post_id}'
            })

    return posts


def get_post_detail(post_url: str) -> Dict:
    """获取单个帖子的详细信息"""
    try:
        temp_file = "temp_post_detail.json"
        cmd = f'autocli read "{post_url}" -f json -o {temp_file}'
        result = subprocess.run(cmd, shell=True, capture_output=True, timeout=30)

        if result.returncode == 0 and os.path.exists(temp_file):
            with open(temp_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            content = data.get('content', '')
            byline = data.get('byline', '')

            # 提取用户ID
            user_id = byline.strip() if byline else 'Unknown'

            # 提取内容摘要
            content_text = re.sub(r'<[^>]+>', '', content)
            content_text = re.sub(r'\s+', ' ', content_text).strip()
            content_summary = content_text[:500] if content_text else ''

            # 清理临时文件
            os.remove(temp_file)

            return {
                'user_id': user_id,
                'content': content_summary
            }
    except Exception as e:
        pass

    return {'user_id': 'Unknown', 'content': ''}


def scrape_all_pages(base_url: str, total_pages: int) -> List[Post]:
    """抓取所有页面"""
    all_posts = []

    for page in range(1, total_pages + 1):
        print(f"\n{'='*60}")
        print(f"Scraping page {page}/{total_pages}...")
        print(f"{'='*60}")

        # 构建URL（如果是第一页，不需要page参数）
        if page == 1:
            url = base_url
        else:
            url = f"{base_url}&page={page}"

        # 抓取页面
        temp_file = f"honor_data/page{page}_temp.json"
        cmd = f'autocli read "{url}" -f json -o {temp_file}'
        result = subprocess.run(cmd, shell=True, capture_output=True, timeout=60)

        if result.returncode == 0 and os.path.exists(temp_file):
            with open(temp_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            posts = parse_page_from_autocli(data.get('content', ''), page)
            print(f"Found {len(posts)} posts on page {page}")

            # 获取每个帖子的详情
            for post in posts:
                print(f"  Getting detail for: {post['title'][:40]}...")

                detail = get_post_detail(post['url'])
                time.sleep(0.5)  # 避免请求过快

                post_obj = Post(
                    title=post['title'],
                    user_id=detail.get('user_id', 'Unknown'),
                    content=detail.get('content', ''),
                    post_url=post['url'],
                    vote_count=post['vote_count'],
                    page_number=page
                )
                all_posts.append(post_obj)

            # 保存临时文件
            save_file = f"honor_data/page{page}_posts.json"
            with open(save_file, 'w', encoding='utf-8') as f:
                json.dump(posts, f, ensure_ascii=False, indent=2)

            os.remove(temp_file)
        else:
            print(f"Failed to scrape page {page}")

        time.sleep(1)  # 页面间隔

    return all_posts


def save_json(posts: List[Post], filepath: Path):
    """保存为JSON"""
    data = [{
        'user_id': p.user_id,
        'product_name': p.product_name,
        'title': p.title,
        'content': p.content,
        'post_time': p.post_time,
        'view_count': p.view_count,
        'comment_count': p.comment_count,
        'post_url': p.post_url,
        'page_number': p.page_number,
        'vote_count': p.vote_count,
        'comments': p.comments
    } for p in posts]

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON saved: {filepath}")


def save_xlsx(posts: List[Post], filepath: Path):
    """保存为XLSX"""
    if not OPENPYXL_AVAILABLE:
        print("openpyxl not available, skipping XLSX")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Honor WIN Posts"

    # 表头
    headers = ['No', 'User ID', 'Product', 'Title', 'Content', 'Post Time',
               'Views', 'Comments', 'URL', 'Page', 'Votes', 'Comments Detail']
    ws.append(headers)

    # 样式
    header_fill = openpyxl.styles.PatternFill(start_color="CF1F25", end_color="CF1F25", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据
    for i, post in enumerate(posts, 1):
        comments_json = json.dumps(post.comments, ensure_ascii=False)
        row = [i, post.user_id, post.product_name, post.title, post.content,
               post.post_time, post.view_count, post.comment_count,
               post.post_url, post.page_number, post.vote_count, comments_json]
        ws.append(row)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i + 1, column=col)
            cell.border = thin_border

    # 列宽
    widths = [8, 20, 15, 40, 50, 15, 10, 10, 50, 8, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(filepath)
    print(f"XLSX saved: {filepath}")


def save_html(posts: List[Post], filepath: Path):
    """保存为HTML"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>Honor WIN Series - Game Experience Feedback</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #cf1f25; text-align: center; margin-bottom: 10px; }}
        .subtitle {{ text-align: center; color: #666; margin-bottom: 30px; }}
        .stats {{ display: flex; justify-content: space-around; margin-bottom: 30px; padding: 20px; background: #f9f9f9; border-radius: 8px; }}
        .stat-item {{ text-align: center; }}
        .stat-value {{ font-size: 28px; font-weight: bold; color: #cf1f25; }}
        .stat-label {{ color: #666; font-size: 14px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 12px 8px; text-align: left; }}
        th {{ background: #cf1f25; color: white; font-weight: 600; position: sticky; top: 0; }}
        tr:nth-child(even) {{ background: #f9f9f9; }}
        tr:hover {{ background: #f0f0f0; }}
        .title-cell {{ max-width: 300px; word-wrap: break-word; }}
        .votes {{ text-align: center; font-weight: bold; color: #cf1f25; }}
        .link-cell a {{ color: #0066cc; text-decoration: none; }}
        .link-cell a:hover {{ text-decoration: underline; }}
        .timestamp {{ text-align: center; color: #999; margin-top: 20px; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Honor WIN Series - Game Experience Feedback</h1>
        <p class="subtitle">User suggestions for Honor WIN Series</p>
        <div class="stats">
            <div class="stat-item"><div class="stat-value">{len(posts)}</div><div class="stat-label">Total Posts</div></div>
            <div class="stat-item"><div class="stat-value">{sum(p.vote_count for p in posts)}</div><div class="stat-label">Total Votes</div></div>
            <div class="stat-item"><div class="stat-value">{sum(p.comment_count for p in posts)}</div><div class="stat-label">Total Comments</div></div>
            <div class="stat-item"><div class="stat-value">{len(set(p.user_id for p in posts))}</div><div class="stat-label">Unique Users</div></div>
        </div>
        <table>
            <thead>
                <tr><th>No</th><th>User ID</th><th>Product</th><th>Title</th><th>Votes</th><th>Comments</th><th>Page</th><th>URL</th></tr>
            </thead>
            <tbody>
"""

    for i, post in enumerate(posts, 1):
        html += f"""                <tr>
                    <td style="text-align:center">{i}</td>
                    <td>{post.user_id}</td>
                    <td>{post.product_name}</td>
                    <td class="title-cell">{post.title}</td>
                    <td class="votes">{post.vote_count}</td>
                    <td style="text-align:center">{post.comment_count}</td>
                    <td style="text-align:center">{post.page_number}</td>
                    <td class="link-cell"><a href="{post.post_url}" target="_blank">View</a></td>
                </tr>
"""

    html += f"""            </tbody>
        </table>
        <p class="timestamp">Generated at: {timestamp}</p>
    </div>
</body>
</html>"""

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"HTML saved: {filepath}")


def main():
    print("=" * 60)
    print("Honor WIN Series - Complete Data Collection")
    print("=" * 60)

    output_dir = Path("honor_data")
    output_dir.mkdir(exist_ok=True)

    # 基础URL
    base_url = "https://club.honor.com/cn//cn//opinion_thread-list.html?filter=lastpost&type=23"

    # 从已保存的第一页数据开始
    all_posts = []

    # 解析第一页
    page1_file = "honor_data/page1_autocli.json"
    if os.path.exists(page1_file):
        with open(page1_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        posts = parse_page_from_autocli(data.get('content', ''), 1)
        print(f"\nPage 1: Found {len(posts)} posts")

        for post in posts:
            detail = get_post_detail(post['url'])
            time.sleep(0.5)

            post_obj = Post(
                title=post['title'],
                user_id=detail.get('user_id', 'Unknown'),
                content=detail.get('content', ''),
                post_url=post['url'],
                vote_count=post['vote_count'],
                page_number=1
            )
            all_posts.append(post_obj)

    print(f"\nTotal posts collected: {len(all_posts)}")

    # 保存数据
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    save_json(all_posts, output_dir / f"honor_win_all_{timestamp}.json")
    save_xlsx(all_posts, output_dir / f"honor_win_all_{timestamp}.xlsx")
    save_html(all_posts, output_dir / f"honor_win_all_{timestamp}.html")

    print(f"\nDone! Data saved to: {output_dir.absolute()}")


if __name__ == "__main__":
    main()