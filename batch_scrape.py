#!/usr/bin/env python3
"""
荣耀建议广场数据批量抓取脚本
抓取所有页面并保存为JSON、XLSX和HTML格式
"""

import json
import re
import os
import time
import subprocess
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict, field

# 导入Excel支持
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

@dataclass
class Comment:
    """评论数据模型"""
    user_id: str = ""
    content: str = ""
    post_time: str = ""
    vote_count: int = 0

@dataclass
class Post:
    """帖子数据模型"""
    user_id: str = ""
    product_name: str = "荣耀WIN系列"
    title: str = ""
    content: str = ""
    post_time: str = ""
    view_count: int = 0
    comment_count: int = 0
    post_url: str = ""
    page_number: int = 1
    vote_count: int = 0  # 建议投票参与人数
    comments: List[Dict] = field(default_factory=list)
    image_urls: List[str] = field(default_factory=list)


def parse_page_structure(raw_output: str, page_num: int) -> List[Post]:
    """从firecrawl输出的原始文本解析帖子数据"""
    posts = []

    # 按换行符分割
    lines = raw_output.split('\n')

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 检测帖子标题（后面有支持/反对）
        title_match = re.match(r'^- link "([^"]+)" \[ref=e(\d+)\]$', line)
        if title_match and i+1 < len(lines) and '支持' in lines[i+1]:
            title = title_match.group(1)

            # 查找用户ID和评论数
            user_id = ''
            comment_count = 0

            # 在接下来几行找用户ID（非数字的link）和评论数（纯数字的link）
            for j in range(i+1, min(i+20, len(lines))):
                next_line = lines[j].strip()

                # 用户ID: link "用户名" [ref=eXXX]，用户名不是纯数字
                user_match = re.match(r'^- link "([^"]+)" \[ref=e(\d+)\]$', next_line)
                if user_match:
                    potential_text = user_match.group(1)
                    # 如果不是纯数字，认为是用户ID
                    if potential_text and not potential_text.isdigit():
                        user_id = potential_text
                        # 在此之后找评论数
                        for k in range(j+1, min(j+5, len(lines))):
                            count_match = re.match(r'^- link "(\d+)" \[ref=e(\d+)\]$', lines[k].strip())
                            if count_match:
                                comment_count = int(count_match.group(1))
                                break
                        break

            post = Post(
                title=title.strip(),
                user_id=user_id.strip(),
                comment_count=comment_count,
                page_number=page_num
            )
            posts.append(post)

        i += 1

    return posts


def scrape_page(page_num: int, base_scrape_id: str) -> Optional[str]:
    """使用firecrawl interact抓取指定页面"""
    try:
        # 首先点击翻页
        if page_num > 1:
            # 点击页码按钮
            click_cmd = f'firecrawl interact -p "点击第{page_num}页的分页按钮" --scrape-id {base_scrape_id}'
            subprocess.run(click_cmd, shell=True, capture_output=True, timeout=60)
            time.sleep(3)

        # 重新抓取当前页面
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f".firecrawl/page_{page_num}_{timestamp}.json"

        interact_cmd = f'firecrawl interact -p "输出当前页面的DOM结构" --scrape-id {base_scrape_id} -o {output_file} --json'
        result = subprocess.run(interact_cmd, shell=True, capture_output=True, text=True, timeout=120)

        return output_file if os.path.exists(output_file) else None

    except Exception as e:
        print(f"  Error scraping page {page_num}: {e}")
        return None


def save_json(posts: List[Post], filepath: Path):
    """保存为JSON格式"""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump([asdict(p) for p in posts], f, ensure_ascii=False, indent=2)
    print(f"  JSON saved: {filepath}")


def save_xlsx(posts: List[Post], filepath: Path):
    """保存为XLSX格式"""
    if not OPENPYXL_AVAILABLE:
        print("  openpyxl not installed, skipping XLSX generation")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Honor WIN Data"

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="CF1F25", end_color="CF1F25", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        "No", "User ID", "Product", "Title", "Content",
        "Post Time", "Views", "Comments", "URL",
        "Page", "Votes", "Comments Detail", "Images"
    ]
    ws.append(headers)

    # 表头样式
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    for i, post in enumerate(posts, 1):
        comments_json = json.dumps(post.comments, ensure_ascii=False)
        images_json = json.dumps(post.image_urls, ensure_ascii=False)

        row = [
            i,
            post.user_id,
            post.product_name,
            post.title,
            post.content,
            post.post_time,
            post.view_count,
            post.comment_count,
            post.post_url,
            post.page_number,
            post.vote_count,
            comments_json,
            images_json
        ]
        ws.append(row)

        # 数据行样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i + 1, column=col)
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal='center')

    # 列宽
    column_widths = [8, 20, 15, 40, 50, 15, 10, 10, 50, 8, 12, 60, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    wb.save(filepath)
    print(f"  XLSX saved: {filepath}")


def save_html(posts: List[Post], filepath: Path):
    """保存为HTML格式"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>Honor WIN Series Feedback Data</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            margin: 0; padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px; margin: 0 auto;
            background: white; padding: 30px;
            border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{ color: #cf1f25; text-align: center; margin-bottom: 10px; }}
        .subtitle {{ text-align: center; color: #666; margin-bottom: 30px; }}
        .stats {{
            display: flex; justify-content: space-around;
            margin-bottom: 30px; padding: 20px;
            background: #f9f9f9; border-radius: 8px;
        }}
        .stat-item {{ text-align: center; }}
        .stat-value {{ font-size: 32px; font-weight: bold; color: #cf1f25; }}
        .stat-label {{ color: #666; font-size: 14px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 12px 8px; text-align: left; }}
        th {{ background-color: #cf1f25; color: white; font-weight: 600; position: sticky; top: 0; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        tr:hover {{ background-color: #f0f0f0; }}
        .title-cell {{ max-width: 300px; word-wrap: break-word; }}
        .comment-count {{ text-align: center; font-weight: bold; }}
        .view-count {{ text-align: center; color: #666; }}
        .page-num {{ text-align: center; }}
        .link-cell a {{ color: #0066cc; text-decoration: none; }}
        .link-cell a:hover {{ text-decoration: underline; }}
        .timestamp {{ text-align: center; color: #999; margin-top: 20px; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Honor Suggestion Square - Game Experience - WIN Series</h1>
        <p class="subtitle">User feedback data for Honor WIN Series</p>
        <div class="stats">
            <div class="stat-item"><div class="stat-value">{len(posts)}</div><div class="stat-label">Total Posts</div></div>
            <div class="stat-item"><div class="stat-value">{sum(p.view_count for p in posts)}</div><div class="stat-label">Total Views</div></div>
            <div class="stat-item"><div class="stat-value">{sum(p.comment_count for p in posts)}</div><div class="stat-label">Total Comments</div></div>
            <div class="stat-item"><div class="stat-value">{len(set(p.user_id for p in posts))}</div><div class="stat-label">Unique Users</div></div>
        </div>
        <table>
            <thead>
                <tr>
                    <th>No</th><th>User ID</th><th>Product</th><th>Title</th>
                    <th>Post Time</th><th>Views</th><th>Comments</th><th>Page</th><th>URL</th>
                </tr>
            </thead>
            <tbody>
"""

    for i, post in enumerate(posts, 1):
        link = f'<a href="{post.post_url}" target="_blank">View</a>' if post.post_url else '-'
        html += f"""                <tr>
                    <td style="text-align:center">{i}</td>
                    <td>{post.user_id}</td>
                    <td>{post.product_name}</td>
                    <td class="title-cell">{post.title}</td>
                    <td>{post.post_time or '-'}</td>
                    <td class="view-count">{post.view_count}</td>
                    <td class="comment-count">{post.comment_count}</td>
                    <td class="page-num">{post.page_number}</td>
                    <td class="link-cell">{link}</td>
                </tr>
"""

    html += f"""            </tbody>
        </table>
        <p class="timestamp">Data captured at: {timestamp}</p>
    </div>
</body>
</html>"""

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  HTML saved: {filepath}")


def main():
    print("=" * 60)
    print("Honor Suggestion Square - Game Experience - WIN Series")
    print("=" * 60)

    # 创建输出目录
    output_dir = Path("honor_data")
    output_dir.mkdir(exist_ok=True)

    # 解析已保存的第一页数据
    raw_file = Path("honor_data/page1_raw.json")
    all_posts = []

    if raw_file.exists():
        with open(raw_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        result = data.get('result', '')
        posts = parse_page_structure(result, 1)
        all_posts.extend(posts)
        print(f"\nPage 1: {len(posts)} posts parsed")

    # 继续抓取第2-19页
    # 由于firecrawl interact不太稳定，我们使用已抓取的数据
    print("\nNote: To get all 19 pages, need to run firecrawl interact for each page")
    print("Currently only page 1 data is available")

    # 如果有更多页面数据，继续解析
    for page in range(2, 20):
        page_file = Path(f"honor_data/page{page}_raw.json")
        if page_file.exists():
            with open(page_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            result = data.get('result', '')
            posts = parse_page_structure(result, page)
            all_posts.extend(posts)
            print(f"Page {page}: {len(posts)} posts parsed")
        else:
            print(f"Page {page}: file not found")

    print(f"\nTotal posts: {len(all_posts)}")

    # 保存数据
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    save_json(all_posts, output_dir / f"honor_win_posts_{timestamp}.json")
    save_xlsx(all_posts, output_dir / f"honor_win_posts_{timestamp}.xlsx")
    save_html(all_posts, output_dir / f"honor_win_posts_{timestamp}.html")

    print(f"\nDone! Output directory: {output_dir.absolute()}")


if __name__ == "__main__":
    main()