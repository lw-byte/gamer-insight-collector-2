#!/usr/bin/env python3
"""
荣耀俱乐部 WIN系列帖子分类器
从功能特性和游戏两个角度对帖子进行分类
"""

import json
import re
from pathlib import Path

def classify_posts(threads):
    """对帖子进行功能特性和游戏分类"""

    # 功能特性分类
    feature_categories = {
        '帧率适配': ['帧率', '165帧', '185帧', '144帧', '120帧', '帧率增强', '插帧', '原生帧'],
        '刷新率档位': ['144Hz', '165Hz', '185Hz', '刷新率', '高刷', '画面撕裂'],
        '画质优化': ['画质', '抗锯齿', '超分', '超分辨率', '分辨率', '高清画质', '精致画质', '画质增强'],
        '幻影稳帧': ['幻影稳帧', '稳帧', '帧率稳定'],
        '极客中心': ['极客中心', '极客模式', '性能模式', '火力全开', '调度', '主频', 'CPU频率'],
        '应用分身': ['分身', '双开', '应用分身', '游戏分身'],
        '游戏适配': ['适配', '游戏适配', '新游适配', '优化'],
        '温控续航': ['温控', '续航', '发热', '降亮度', '电池', '功耗', '耗电'],
        '外设支持': ['手柄', '肩键', '外设', '游戏手柄'],
        '其他功能': []
    }

    # 游戏分类
    game_categories = {
        '三角洲行动': ['三角洲', '三角洲行动'],
        '无畏契约': ['无畏契约', '瓦', '瓦手', 'VALORANT'],
        '王者荣耀': ['王者荣耀', '王者', '王者世界'],
        '英雄联盟': ['英雄联盟', 'LOL', '英雄联盟手游'],
        '原神': ['原神'],
        '明日方舟': ['明日方舟', '方舟', '终末地'],
        '妄想山海': ['妄想山海'],
        '逆水寒': ['逆水寒'],
        '和平精英': ['和平精英', '吃鸡'],
        '鸣潮': ['鸣潮'],
        '晶核': ['晶核'],
        '异环': ['异环'],
        '黑神话': ['黑神话'],
        '我的世界': ['我的世界', 'MC'],
        'PC游戏': ['PC端', '电脑端', '模拟器'],
        '其他游戏': []
    }

    classified = []

    for thread in threads:
        content = thread.get('content', '')
        title = content.split()[0] if content else ''
        body = ' '.join(content.split()[1:]) if len(content.split()) > 1 else ''

        # 功能特性分类
        features = []
        for category, keywords in feature_categories.items():
            if category == '其他功能':
                continue
            if any(kw in content for kw in keywords):
                features.append(category)
                break

        if not features:
            features = ['其他功能']

        # 游戏分类
        games = []
        for game, keywords in game_categories.items():
            if game == '其他游戏':
                continue
            if any(kw in content for kw in keywords):
                games.append(game)
                break

        if not games:
            games = ['其他游戏']

        classified.append({
            **thread,
            'title': title,
            'body': body,
            'feature_category': features[0],
            'feature_categories': features,
            'game_category': games[0],
            'game_categories': games
        })

    return classified

def analyze_classification(classified):
    """统计分析分类结果"""
    # 功能特性统计
    feature_stats = {}
    for t in classified:
        cat = t['feature_category']
        feature_stats[cat] = feature_stats.get(cat, 0) + 1

    # 游戏统计
    game_stats = {}
    for t in classified:
        cat = t['game_category']
        game_stats[cat] = game_stats.get(cat, 0) + 1

    return feature_stats, game_stats

def generate_report(classified, feature_stats, game_stats):
    """生成分析报告"""
    report = f"""# 荣耀WIN系列用户反馈分析报告

## 概述
- 总帖子数: {len(classified)}
- 抓取时间: 自动生成

---

## 一、功能特性维度分析

### 1.1 功能分类统计

| 功能类别 | 帖子数量 | 占比 |
|---------|---------|------|
"""
    for cat, count in sorted(feature_stats.items(), key=lambda x: -x[1]):
        pct = count / len(classified) * 100
        report += f"| {cat} | {count} | {pct:.1f}% |\n"

    report += """
### 1.2 功能类别说明

| 类别 | 说明 | 典型需求 |
|------|------|---------|
| 帧率适配 | 游戏帧率支持 | 165帧、185帧适配 |
| 刷新率档位 | 屏幕刷新率档位 | 增加144Hz/165Hz档位 |
| 画质优化 | 画面质量提升 | 抗锯齿、超分、画质增强 |
| 幻影稳帧 | 帧率稳定技术 | 新游戏适配 |
| 极客中心 | 性能调度控制 | 火力全开、频率解锁 |
| 应用分身 | 应用多开 | 游戏多开 |
| 游戏适配 | 游戏优化 | 新游戏提前适配 |
| 温控续航 | 功耗发热 | 续航优化 |
| 外设支持 | 外部设备 | 游戏手柄、肩键 |
"""

    report += """

---

## 二、游戏维度分析

### 2.1 游戏分类统计

| 游戏名称 | 帖子数量 | 占比 |
|---------|---------|------|
"""
    for game, count in sorted(game_stats.items(), key=lambda x: -x[1]):
        pct = count / len(classified) * 100
        report += f"| {game} | {count} | {pct:.1f}% |\n"

    report += """
### 2.2 高热度游戏说明

"""

    # 找出热门游戏
    hot_games = sorted(game_stats.items(), key=lambda x: -x[1])[:5]
    for game, count in hot_games:
        if game == '其他游戏':
            continue
        # 找出该游戏的典型帖子
        game_posts = [t for t in classified if game in t['game_category']][:3]
        report += f"""#### {game} ({count}条)

"""
        for post in game_posts:
            report += f"- **{post['title']}**: {post['body'][:60]}...\n"
        report += "\n"

    report += """---

## 三、关键洞察

### 3.1 功能痛点TOP3
"""
    top_features = sorted(feature_stats.items(), key=lambda x: -x[1])[:3]
    for i, (cat, count) in enumerate(top_features, 1):
        report += f"{i}. **{cat}** - {count}条反馈\n"

    report += """
### 3.2 热门游戏TOP5
"""
    top_games = sorted(game_stats.items(), key=lambda x: -x[1])[:5]
    for i, (game, count) in enumerate(top_games, 1):
        if game != '其他游戏':
            report += f"{i}. **{game}** - {count}条反馈\n"

    report += """
### 3.3 典型用户诉求

1. **帧率与画质不可兼得**: 用户希望在165帧下也能保持高清画质
2. **WIN RT差异化对待**: WIN有插帧功能但WIN RT没有，用户不满
3. **极客中心下放**: 希望将Magic系列平板的极客中心功能下放到WIN系列
4. **新游戏预适配**: 希望官方能提前适配热门新游如《异环》

---

*报告自动生成 - Honor WIN Insights Collector*
"""

    return report

def save_classified_data(classified, output_dir='.'):
    """保存分类后的数据"""
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    # 保存JSON
    with open(output_dir / 'classified_posts.json', 'w', encoding='utf-8') as f:
        json.dump(classified, f, ensure_ascii=False, indent=2)

    # 保存Excel
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "分类帖子"

    headers = ['序号', '页号', '帖子ID', '标题', '内容', '功能分类', '游戏分类',
               '参与人数', '评论数', '帖子链接']
    ws.append(headers)

    for i, post in enumerate(classified, 1):
        row = [
            i,
            f"第{post['page']}页",
            post['thread_id'],
            post['title'],
            post['body'][:200],
            post['feature_category'],
            post['game_category'],
            post['participants'],
            post['comments'],
            post['url']
        ]
        ws.append(row)

    wb.save(output_dir / 'classified_posts.xlsx')

    # 保存按功能分类的Excel
    wb2 = Workbook()
    for feature in sorted(set(p['feature_category'] for p in classified)):
        if feature == '其他功能':
            continue
        ws2 = wb2.create_sheet(title=feature[:20])
        ws2.append(['序号', '页号', '帖子ID', '标题', '内容', '游戏分类', '参与人数', '链接'])

        posts = [p for p in classified if p['feature_category'] == feature]
        for i, p in enumerate(posts, 1):
            ws2.append([i, f"第{p['page']}页", p['thread_id'], p['title'],
                       p['body'][:100], p['game_category'], p['participants'], p['url']])

    wb2.save(output_dir / 'classified_by_feature.xlsx')

    # 保存按游戏分类的Excel
    wb3 = Workbook()
    for game in sorted(set(p['game_category'] for p in classified)):
        if game == '其他游戏':
            continue
        ws3 = wb3.create_sheet(title=game[:20])
        ws3.append(['序号', '页号', '帖子ID', '标题', '内容', '功能分类', '参与人数', '链接'])

        posts = [p for p in classified if p['game_category'] == game]
        for i, p in enumerate(posts, 1):
            ws3.append([i, f"第{p['page']}页", p['thread_id'], p['title'],
                       p['body'][:100], p['feature_category'], p['participants'], p['url']])

    wb3.save(output_dir / 'classified_by_game.xlsx')

    print(f"分类数据已保存:")
    print(f"  - {output_dir / 'classified_posts.json'}")
    print(f"  - {output_dir / 'classified_posts.xlsx'}")
    print(f"  - {output_dir / 'classified_by_feature.xlsx'}")
    print(f"  - {output_dir / 'classified_by_game.xlsx'}")

def main():
    # 读取原始数据
    data_file = Path('honor_club_data.json')
    if not data_file.exists():
        print("未找到数据文件 honor_club_data.json")
        return

    with open(data_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    threads = data.get('all_threads', [])

    # 分类
    classified = classify_posts(threads)

    # 统计分析
    feature_stats, game_stats = analyze_classification(classified)

    # 生成报告
    report = generate_report(classified, feature_stats, game_stats)

    # 保存报告
    with open('classification_report.md', 'w', encoding='utf-8') as f:
        f.write(report)

    print("\n分析报告已生成: classification_report.md")

    # 保存分类数据
    save_classified_data(classified)

    # 打印统计摘要
    print("\n" + "=" * 50)
    print("分类统计摘要")
    print("=" * 50)

    print("\n功能特性分类:")
    for cat, count in sorted(feature_stats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}条")

    print("\n游戏分类:")
    for game, count in sorted(game_stats.items(), key=lambda x: -x[1])[:10]:
        print(f"  {game}: {count}条")

if __name__ == '__main__':
    main()